# -*- coding: utf-8 -*-
"""
Created on Wed Jun 24 22:22:32 2020

@author: erich
"""

import openpyxl, os

os.chdir(r'C:\Users\erich\Desktop\Python')

# READING SPREADSHEET
# Create link with workbook within Python
workbook1 = openpyxl.load_workbook(r'Spicy Crispy Chicken.xlsx')
print(type(workbook1))

# Create sheet object within Python
sheet1 = workbook1.get_sheet_by_name('Sheet1')
print(type(sheet1))

# Call all sheets in a workbook
sheetsList = workbook1.get_sheet_names()
print(sheetsList)

# Refer to a cell within the sheet
cell1 = sheet1['A1']
print(cell1)

# Call cell's value using A1 syntax
cellValue1 = cell1.value #Could also do cellValue1=sheet1['A1'].value
print(cellValue1)

# Call cell's value using row/column #
cell2 = sheet1.cell(row=3, column=3)
cellValue2 = cell2.value
print(cellValue2)

# Using for loops to call a portion of the workbook
for i in range(2,10):
    for j in range(2,10):
        print(i,sheet1.cell(row=i, column=j).value)
        
print(r'_______________________')

# CREATING A NEW WORKBOOK
# Create a new workbook
wb1 = openpyxl.Workbook()
print(wb1.get_sheet_names())
newSheet1 = wb1.get_sheet_by_name('Sheet')

# Write a value into the cell
newSheet1['A1']= 43
newSheet1['A2']= 'Hello'
print(str(newSheet1['A1'].value)+'\n'+newSheet1['A2'].value)

# Create new sheet and order it
newSheet2 = wb1.create_sheet(index=0)
print(wb1.get_sheet_names())

# Rename sheets
print(newSheet2.title)
newSheet2.title = 'newSheet2'
print(newSheet2.title)
print(wb1.get_sheet_names())

#wb1.save(r'openpyxlSample.xlsx')